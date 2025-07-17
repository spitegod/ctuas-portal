
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from main.models import Teacher, FirstSem, SecondSem, MethodicalWork, OrgMethodicalWork, SciResearchWork, ContractWork, SciMethodicalWork, PublicWork, Remark
from django.db.models import Q
from django.shortcuts import render, redirect
from main.models import Teacher, EducationalMethodicalWork, OrganizationalMethodicalWork,ResearchWork
from .models import Teacher, EducationalMethodicalWork, OrganizationalMethodicalWork, ResearchWork, ContractResearchWork,ScientificMethodicalWork,SocialEducationalWork,TeacherRemark, QualificationUpgrade, QualificationUpgrade, PublishedSciWork, Raising, Recommendation,TeacherPermission, TeachingLoad
from .forms import EducationalMethodicalWorkForm, OrganizationalMethodicalWorkForm, ResearchWorkForm, ContractResearchWorkForm
import openpyxl
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from .utils.user_actions import handle_add_user, handle_delete_user
from django.urls import reverse
from django.http import HttpResponseForbidden
from .utils.user_actions import handle_update_permissions
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)  # авторизуем независимо от прав
            return redirect('dashboard')  # там разруливаем, куда перекинуть
        else:
            return render(request, 'main/login.html', {'error': 'Неверный логин или пароль'})
    return render(request, 'main/login.html')

def index(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    else:
        return redirect('login')


def logout_view(request):
    logout(request)
    return redirect('login')


def dashboard(request):
    tab = request.GET.get('tab') or request.POST.get('active_tab') or '1'

    if not request.user.is_authenticated:
        return redirect('login')

    if request.method == 'POST':
        action = request.POST.get("action")

        if action == "update_permissions":
            user_id = request.POST.get("user_id")
            handle_update_permissions(request, user_id)
            return redirect(f"{reverse('dashboard')}?tab=13")

        elif action == 'select_ip_teacher':
            selected_teacher_id = request.POST.get('selected_teacher_id')
            print("Получен selected_teacher_id из POST:", selected_teacher_id)
            if selected_teacher_id:
                request.session['selected_teacher_id'] = selected_teacher_id
            else:
                request.session.pop('selected_teacher_id', None)
            return redirect('dashboard')

        elif request.user.is_superuser:
            if action == 'add':
                handle_add_user(request)
                return redirect(f"{reverse('dashboard')}?tab={tab}")
            elif action == 'delete':
                handle_delete_user(request)
                return redirect(f"{reverse('dashboard')}?tab={tab}")

        else:
            teacher = Teacher.objects.filter(user=request.user).first()

            if action == 'add_user_from_teacher':
                handle_add_user(request)
                return redirect(f"{reverse('dashboard')}?tab=12")

            elif action == 'delete_user_from_teacher':
                handle_delete_user(request, teacher)
                return redirect(f"{reverse('dashboard')}?tab=12")

            elif action == 'add_teacher':
                full_name = request.POST.get('full_name')
                teacher_type = request.POST.get('teacher_type')
                position = request.POST.get('position')
                rate = request.POST.get('rate')
                degree = request.POST.get('degree')
                rank = request.POST.get('rank')
                email = request.POST.get('email')
                phone = request.POST.get('phone')
                birthday = request.POST.get('birthday') or None
                address = request.POST.get('address')

                if not full_name or not teacher_type or not position or not rate:
                    messages.error(request, "Поля ФИО, Тип, Должность и Ставка обязательны.")
                else:
                    Teacher.objects.create(
                        full_name=full_name,
                        teacher_type=teacher_type,
                        position=position,
                        rate=rate,
                        degree=degree or None,
                        rank=rank or None,
                        email=email or None,
                        phone=phone or None,
                        birthday=birthday,
                        address=address or None
                    )
                    messages.success(request, f"Преподаватель '{full_name}' успешно добавлен.")

            elif action == 'delete_teacher':
                teacher_id = request.POST.get('delete_teacher')
                try:
                    teacher = Teacher.objects.get(id=teacher_id)
                    name = teacher.full_name
                    if teacher.user:
                        teacher.user = None
                        teacher.save()
                    teacher.delete()
                    messages.success(request, f"Преподаватель '{name}' удалён.")
                except Teacher.DoesNotExist:
                    messages.error(request, "Преподаватель не найден.")

    if request.user.is_superuser:
        staff_users = User.objects.filter(Q(is_staff=True) | Q(teacher__isnull=False)).distinct()
        teachers = Teacher.objects.all()

        selected_teacher = None
        selected_teacher_id = request.session.get('selected_teacher_id')
        print("ID из сессии:", selected_teacher_id)
        if selected_teacher_id:
            try:
                selected_teacher = Teacher.objects.get(id=selected_teacher_id)
                print("Найден преподаватель:", selected_teacher.full_name)
            except Teacher.DoesNotExist:
                print("Преподаватель не найден")
                request.session.pop('selected_teacher_id', None)

        teachingload = TeachingLoad.objects.filter(teacher=selected_teacher) if selected_teacher else TeachingLoad.objects.none()
        print("TeachingLoad найдено записей:", teachingload.count())

        edu_methodwork = EducationalMethodicalWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        org_methodwork = OrganizationalMethodicalWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        sci_researchwork = ResearchWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        contractwork = ContractResearchWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        sci_methodwork = ScientificMethodicalWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        published_sciwork = PublishedSciWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        social_eduwork = SocialEducationalWork.objects.filter(teacher=selected_teacher) if selected_teacher else []
        remark = TeacherRemark.objects.filter(teacher=selected_teacher) if selected_teacher else []
        raising = Raising.objects.filter(teacher=selected_teacher) if selected_teacher else []
        recommendation = Recommendation.objects.filter(teacher=selected_teacher) if selected_teacher else []

        return render(request, 'main/admin_dashboard.html', {
            'active_tab': int(tab),
            'staff_users': staff_users,
            'teachers': teachers,
            'teachingload': teachingload,
            'edu_methodwork': edu_methodwork,
            'org_methodwork': org_methodwork,
            'sci_researchwork': sci_researchwork,
            'contractwork': contractwork,
            'sci_methodwork': sci_methodwork,
            'published_sciwork': published_sciwork,
            'social_eduwork': social_eduwork,
            'remark': remark,
            'raising': raising,
            'recommendation': recommendation,
            'selected_teacher': selected_teacher,
        })

    elif request.user.is_staff and not request.user.is_superuser:
        return teacher_dashboard(request, tab=tab)

    return HttpResponseForbidden("Доступ запрещён.")

    
def teacher_dashboard(request,tab="1"):
    teacher = request.user.teacher
    permissions = getattr(teacher, 'permissions', None)
    print(f"[DEBUG] can_edit_pps = {permissions.can_edit_pps if permissions else 'NO PERMS'}")
    last_initial = teacher.full_name.strip()[0].upper() if teacher.full_name else ''
    active_tab = int(tab)
    # === Выбор учебного года ===
    selected_year = request.GET.get("year") or request.session.get("selected_year")
    if not selected_year:
        current = datetime.now().year
        selected_year = f"{current-1}–{current}"
    request.session["selected_year"] = selected_year

    edit_id_tab2 = edit_id_tab3 = edit_id_tab4 = edit_id_tab5 = edit_id_tab6 = edit_id_tab7 = edit_id_tab8 = edit_id_tab9 = edit_id_tab10 = None
    edit_work_tab2 = edit_work_tab3 = edit_work_tab4 = edit_work_tab5 = edit_work_tab6 = edit_work_tab7 = edit_work_tab8 = edit_work_tab9 = edit_work_tab10 = None
    

    if request.method == 'POST':
        form_type = request.POST.get("form_type")
        action = request.POST.get("action")
        if action == "add_user_from_teacher":
            return handle_add_user(request, teacher)
        elif action == "delete_user_from_teacher":
             return handle_delete_user(request, teacher)

        # === Вкладка 2: Учебно-методическая ===
        if request.POST.get("edit_teaching_method"):
            edit_id_tab2 = request.POST.get("edit_teaching_method")
            edit_work_tab2 = EducationalMethodicalWork.objects.filter(id=edit_id_tab2, teacher=teacher).first()
            active_tab = 2

        elif form_type == "teaching_method":
             title = request.POST.get("title", "").strip()

             if not title:
                messages.error(request, "Поле 'Наименование работы' обязательно для заполнения.")
                return redirect('/dashboard?tab=2')
            
             start_date = request.POST.get("start_date")
             if not start_date:
               messages.error(request, "Поле 'Дата начала' обязательно для заполнения.")
               return redirect('/dashboard?tab=2')
        
             work_id = request.POST.get("work_id")
             academic_year = request.POST.get("academic_year", selected_year)
             if work_id:
                work = EducationalMethodicalWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.title = request.POST.get("title")
                    work.start_date = request.POST.get("start_date")
                    work.end_date = request.POST.get("end_date") or None
                    work.completed = request.POST.get("completed")
                    work.academic_year = academic_year
                    work.save()
                    messages.success(request, "Запись обновлена.")
             else:
                EducationalMethodicalWork.objects.create(
                    teacher=teacher,
                    title=request.POST.get("title"),
                    start_date=request.POST.get("start_date"),
                    end_date=request.POST.get("end_date") or None,
                    completed=request.POST.get("completed"),
                    academic_year=academic_year 
                )
                messages.success(request, "Учебно-методическая работа добавлена.")
             return redirect(f'/dashboard?tab=2')

        elif form_type == "delete_teaching_method":
            EducationalMethodicalWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=2')

        # === Вкладка 3: Организационно-методическая ===
        if request.POST.get("edit_organizational_method"):
            edit_id_tab3 = request.POST.get("edit_organizational_method")
            edit_work_tab3 = OrganizationalMethodicalWork.objects.filter(id=edit_id_tab3, teacher=teacher).first()
            active_tab = 3

        elif form_type == "organizational_method":
            title = request.POST.get("title", "").strip()
            if not title: 
                messages.error(request, "Поле 'Наименование работы' обязательно для заполнения.")
                return redirect(f'/dashboard?tab=3')
            
            start_date = request.POST.get("start_date")
            if not start_date:
               messages.error(request, "Поле 'Дата начала' обязательно для заполнения.")
               return redirect('/dashboard?tab=3')
        
            work_id = request.POST.get("work_id")
            if work_id:
                work = OrganizationalMethodicalWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.title = request.POST.get("title")
                    work.start_date = request.POST.get("start_date")
                    work.end_date = request.POST.get("end_date") or None
                    work.completed = request.POST.get("completed")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                OrganizationalMethodicalWork.objects.create(
                    teacher=teacher,
                    title=request.POST.get("title"),
                    start_date=request.POST.get("start_date"),
                    end_date=request.POST.get("end_date") or None,
                    completed=request.POST.get("completed")
                )
                messages.success(request, "Орг.-методическая работа добавлена.")
            return redirect(f'/dashboard?tab=3')

        elif form_type == "delete_organizational_method":
            OrganizationalMethodicalWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=3')

        # === Вкладка 4: Научно-исследовательская ===
        if request.POST.get("edit_research_work"):
            edit_id_tab4 = request.POST.get("edit_research_work")
            edit_work_tab4 = ResearchWork.objects.filter(id=edit_id_tab4, teacher=teacher).first()
            active_tab = 4

        elif form_type == "research_work":
            topic = request.POST.get("topic", "").strip()
            if not topic:
               messages.error(request, "Поле 'Тема работы' обязательно для заполнения.")
               return redirect(f'/dashboard?tab=4')
           
            start_date = request.POST.get("start_date")
            if not start_date:
               messages.error(request, "Поле 'Дата начала' обязательно для заполнения.")
               return redirect('/dashboard?tab=4')
        
            work_id = request.POST.get("work_id")
            if work_id:
                work = ResearchWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.topic = request.POST.get("topic")
                    work.start_date = request.POST.get("start_date")
                    work.end_date = request.POST.get("end_date") or None
                    work.completed = request.POST.get("completed")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                ResearchWork.objects.create(
                    teacher=teacher,
                    topic=request.POST.get("topic"),
                    start_date=request.POST.get("start_date"),
                    end_date=request.POST.get("end_date") or None,
                    completed=request.POST.get("completed")
                )
                messages.success(request, "Научно-исследовательская работа добавлена.")
            return redirect(f'/dashboard?tab=4')

        elif form_type == "delete_research_work":
            ResearchWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=4')

        # === Вкладка 5: Хоздоговорная НИР ===
        if request.POST.get("edit_contract_research"):
            edit_id_tab5 = request.POST.get("edit_contract_research")
            edit_work_tab5 = ContractResearchWork.objects.filter(id=edit_id_tab5, teacher=teacher).first()
            active_tab = 5

        elif form_type == "contract_research":
            topic = request.POST.get("topic", "").strip()
            position = request.POST.get("position", "").strip()

            if not topic or not position:
             messages.error(request, "Поля 'Тема' и 'Должность' обязательны для заполнения.")
             return redirect(f'/dashboard?tab=5')
         
            start_date = request.POST.get("start_date")
            if not start_date:
               messages.error(request, "Поле 'Дата начала' обязательно для заполнения.")
               return redirect('/dashboard?tab=5')
         
            work_id = request.POST.get("work_id")
            if work_id:
                work = ContractResearchWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.topic = request.POST.get("topic")
                    work.position = request.POST.get("position")
                    work.start_date = request.POST.get("start_date")
                    work.end_date = request.POST.get("end_date") or None
                    work.completed = request.POST.get("completed")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                ContractResearchWork.objects.create(
                    teacher=teacher,
                    topic=request.POST.get("topic"),
                    position=request.POST.get("position"),
                    start_date=request.POST.get("start_date"),
                    end_date=request.POST.get("end_date") or None,
                    completed=request.POST.get("completed")
                )
                messages.success(request, "Хоздоговорная НИР добавлена.")
            return redirect(f'/dashboard?tab=5')

        elif form_type == "delete_contract_research":
            ContractResearchWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=5')

        # === Вкладка 6: Научно-методическая ===
        if request.POST.get("edit_scientific_method"):
            edit_id_tab6 = request.POST.get("edit_scientific_method")
            edit_work_tab6 = ScientificMethodicalWork.objects.filter(id=edit_id_tab6, teacher=teacher).first()
            active_tab = 6

        elif form_type == "scientific_method":
            work_id = request.POST.get("work_id")
            topic = request.POST.get("topic")

            if not topic:
                messages.error(request, "Поле 'Наименование темы' обязательно для заполнения.")
                return redirect(f'/dashboard?tab=6')
            
            start_date = request.POST.get("start_date")
            if not start_date:
               messages.error(request, "Поле 'Дата начала' обязательно для заполнения.")
               return redirect('/dashboard?tab=6')

            if work_id:
                work = ScientificMethodicalWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.topic = topic
                    work.start_date = request.POST.get("start_date")
                    work.end_date = request.POST.get("end_date") or None
                    work.completed = request.POST.get("completed")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                ScientificMethodicalWork.objects.create(
                    teacher=teacher,
                    topic=topic,
                    start_date=request.POST.get("start_date"),
                    end_date=request.POST.get("end_date") or None,
                    completed=request.POST.get("completed")
                )
                messages.success(request, "Научно-методическая работа добавлена.")
            return redirect(f'/dashboard?tab=6')

        elif form_type == "delete_scientific_method":
            ScientificMethodicalWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=6')

        # === Вкладка 7: Публикации ===
        if request.POST.get("edit_published_work"):
            edit_id_tab7 = request.POST.get("edit_published_work")
            edit_work_tab7 = PublishedSciWork.objects.filter(id=edit_id_tab7, teacher=teacher).first()
            active_tab = 7

        elif form_type == "published_work":
            work_id = request.POST.get("work_id")
            title = request.POST.get("title")  # Наименование и вид работы
            output = request.POST.get("output")  # Выходные данные
            size = request.POST.get("size")  # Объем
            autors = request.POST.get("autors")  # Соавторы

            if not title:
                messages.error(request, "Поле 'Наименование и вид работы' обязательно для заполнения.")
                return redirect(f'/dashboard?tab=7')

            if work_id:
                work = PublishedSciWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.title = title
                    work.output = output or ""
                    work.size = size or ""
                    work.autors = autors or ""
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                PublishedSciWork.objects.create(
                    teacher=teacher,
                    title=title,
                    output=output or "",
                    size=size or "",
                    autors=autors or ""
                )
                messages.success(request, "Публикация добавлена.")
            return redirect(f'/dashboard?tab=7')

        elif form_type == "delete_published_work":
            PublishedSciWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=7')

        elif form_type == "import_published_from_library":
            from parsers import import_published_from_library

            result = import_published_from_library(teacher)

            if isinstance(result, tuple) and result[0] == -1:
                _, error_message = result
                messages.error(request, f"Ошибка импорта: {error_message}")
            elif result > 0:
                messages.success(request, f"Импортировано публикаций: {result}")
            else:
                messages.warning(request, "Публикации не найдены или уже добавлены.")
            return redirect(f'/dashboard?tab=7')


        # === Вкладка 8: Общественная работа ===
        if request.POST.get("edit_social_work"):
            edit_id_tab8 = request.POST.get("edit_social_work")
            edit_work_tab8 = SocialEducationalWork.objects.filter(id=edit_id_tab8, teacher=teacher).first()
            active_tab = 8

        elif form_type == "social_work":
            title = request.POST.get("title", "").strip()
            if not title:
             messages.error(request, "Поле 'Наименование работы' обязательно для заполнения.")
             return redirect(f'/dashboard?tab=8')
            work_id = request.POST.get("work_id")
            if work_id:
                work = SocialEducationalWork.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.title = request.POST.get("title")
                    work.completed = request.POST.get("completed")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                SocialEducationalWork.objects.create(
                    teacher=teacher,
                    title=request.POST.get("title"),
                    completed=request.POST.get("completed")
                )
                messages.success(request, "Запись добавлена.")
            return redirect(f'/dashboard?tab=8')

        elif form_type == "delete_social_work":
            SocialEducationalWork.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=8')

        # === Вкладка 9: Замечания ===
        if request.POST.get("edit_remark"):
            edit_id_tab9 = request.POST.get("edit_remark")
            edit_work_tab9 = TeacherRemark.objects.filter(id=edit_id_tab9, teacher=teacher).first()
            active_tab = 9

        elif form_type == "teacher_remark":
            content = request.POST.get("content", "").strip()
            if not content:
               messages.error(request, "Поле 'Содержание замечания' обязательно для заполнения.")
               return redirect(f'/dashboard?tab=9')
           
            work_id = request.POST.get("work_id")
            if work_id:
                work = TeacherRemark.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.date = request.POST.get("date")
                    work.content = request.POST.get("content")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                TeacherRemark.objects.create(
                    teacher=teacher,
                    date=request.POST.get("date"),
                    content=request.POST.get("content")
                )
                messages.success(request, "Замечание добавлено.")
            return redirect(f'/dashboard?tab=9')

        elif form_type == "delete_teacher_remark":
            TeacherRemark.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=9')

        # === Вкладка 10: Повышение квалификации ===
        if request.POST.get("edit_qualification"):
              title = request.POST.get("title", "").strip()
              place = request.POST.get("place", "").strip()
              document_number = request.POST.get("document_number", "").strip()
              date = request.POST.get("date", "").strip()
              duration = request.POST.get("duration", "").strip()
              volume = request.POST.get("volume", "").strip()
              work_id = request.POST.get("work_id")

    # Проверка обязательных полей
              if not title or not place or not document_number or not date or not duration or not volume:
                 messages.error(request, "Все поля: 'Название', 'Место', 'Номер документа', 'Дата', 'Срок прохождения', 'Объём' обязательны для заполнения.")
                 return redirect(f'/dashboard?tab=10')
            
              edit_id_tab10 = request.POST.get("edit_qualification")
              edit_work_tab10 = QualificationUpgrade.objects.filter(id=edit_id_tab10, teacher=teacher).first()
              active_tab = 10

        elif form_type == "qualification_upgrade":
            work_id = request.POST.get("work_id")
            date = request.POST.get("date")
            if not date:
                messages.error(request, "Поле 'Дата' обязательно для заполнения.")
                return redirect(f'/dashboard?tab=10')
    
            if work_id:
                work = QualificationUpgrade.objects.filter(id=work_id, teacher=teacher).first()
                if work:
                    work.title = request.POST.get("title")
                    work.location = request.POST.get("place")
                    work.document_number = request.POST.get("document_number")
                    work.date = request.POST.get("date")
                    work.duration = request.POST.get("duration")
                    work.volume = request.POST.get("volume")
                    work.save()
                    messages.success(request, "Запись обновлена.")
            else:
                QualificationUpgrade.objects.create(
                    teacher=teacher,
                    title=request.POST.get("title"),
                    location=request.POST.get("place"),
                    document_number=request.POST.get("document_number"),
                    date=request.POST.get("date")  or None,
                    duration=request.POST.get("duration"),
                    volume=request.POST.get("volume")
                )
                messages.success(request, "Повышение квалификации добавлено.")
            return redirect(f'/dashboard?tab=10')

        elif form_type == "delete_qualification_upgrade":
            QualificationUpgrade.objects.filter(id=request.POST.get("work_id"), teacher=teacher).delete()
            messages.success(request, "Запись удалена.")
            return redirect(f'/dashboard?tab=10')

    
    context = {
    # Данные по вкладкам
    'works': EducationalMethodicalWork.objects.filter(teacher=teacher, academic_year=selected_year),
    'org_works': OrganizationalMethodicalWork.objects.filter(teacher=teacher),
    'research_works': ResearchWork.objects.filter(teacher=teacher),
    'contract_works': ContractResearchWork.objects.filter(teacher=teacher),
    'scientific_works': ScientificMethodicalWork.objects.filter(teacher=teacher),
    'published_works': PublishedSciWork.objects.filter(teacher=teacher),
    'social_works': SocialEducationalWork.objects.filter(teacher=teacher),
    'teacher_remarks': TeacherRemark.objects.filter(teacher=teacher),
    'qualification_upgrades': QualificationUpgrade.objects.filter(teacher=teacher),

    # Редактируемые записи
    'edit_work_tab2': edit_work_tab2,
    'edit_work_tab3': edit_work_tab3,
    'edit_work_tab4': edit_work_tab4,
    'edit_work_tab5': edit_work_tab5,
    'edit_work_tab6': edit_work_tab6,
    'edit_work_tab7': edit_work_tab7,
    'edit_work_tab8': edit_work_tab8,
    'edit_work_tab9': edit_work_tab9,
    'edit_work_tab10': edit_work_tab10,

    # Основные данные
    'active_tab': active_tab,
    'teacher': teacher,
    'permissions': permissions,
    'last_initial': last_initial,

    # Учебные года
    'selected_year': selected_year,
    'available_years': ["2022–2023", "2023–2024", "2024–2025"],

    # Права доступа
    'can_edit_pps': permissions.can_edit_pps if permissions else False,
    'can_manage_accounts': permissions.can_manage_accounts if permissions else False,
    'can_edit_ip': permissions.can_edit_ip if permissions else False,

    # Для отображения списка сотрудников (если есть доступ)
    'staff_users': User.objects.filter(is_staff=True) if permissions and permissions.can_manage_accounts else [],

    # Преподаватели (для tab11, только если разрешено)
    'teachers': Teacher.objects.all() if permissions and permissions.can_edit_pps else [],
    

    'teachingload': TeachingLoad.objects.filter(teacher=request.user.teacher) if hasattr(request.user, 'teacher') else []
}

    return render(request, 'main/teacher_dashboard.html', context)

def export_qualification_excel(request):
    teacher = request.user.teacher
    qualifications = QualificationUpgrade.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Повышение квалификации"

    headers = ["Название", "Место", "Номер документа", "Дата", "Срок", "Объём"]
    ws.append(headers)

    for q in qualifications:
        ws.append([
            q.title,
            q.location,
            q.document_number,
            q.date.strftime('%Y-%m-%d') if q.date else '',
            q.duration,
            q.volume
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="qualification.xlsx"'
    wb.save(response)
    return response

def export_teaching_excel(request):
    teacher = request.user.teacher
    selected_year = request.GET.get("year") or request.session.get("selected_year")

    if selected_year:
        works = EducationalMethodicalWork.objects.filter(teacher=teacher, academic_year=selected_year)
    else:
        works = EducationalMethodicalWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Учебно-методическая"

    ws.append(["Наименование", "Дата начала", "Дата окончания", "Отметка", "Учебный год"])
    for w in works:
        ws.append([w.title, w.start_date, w.end_date, w.completed, w.academic_year])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="teaching_{selected_year or "all"}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_organizational_excel(request):
    teacher = request.user.teacher
    works = OrganizationalMethodicalWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Орг-методическая"

    ws.append(["Наименование", "Дата начала", "Дата окончания", "Отметка"])
    for w in works:
        ws.append([w.title, w.start_date, w.end_date, w.completed])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="organizational.xlsx"'
    wb.save(response)
    return response


@login_required
def export_organizational_excel(request):
    teacher = request.user.teacher
    works = OrganizationalMethodicalWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Орг-методическая"

    ws.append(["Наименование", "Дата начала", "Дата окончания", "Отметка"])
    for w in works:
        ws.append([w.title, w.start_date, w.end_date, w.completed])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="organizational.xlsx"'
    wb.save(response)
    return response


@login_required
def export_research_excel(request):
    teacher = request.user.teacher
    works = ResearchWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "НИР"

    ws.append(["Тема", "Дата начала", "Дата окончания", "Отметка"])
    for w in works:
        ws.append([w.topic, w.start_date, w.end_date, w.completed])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="research.xlsx"'
    wb.save(response)
    return response

@login_required
def export_contract_excel(request):
    teacher = request.user.teacher
    works = ContractResearchWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Хоздоговорная НИР"

    ws.append(["Тема", "Должность", "Дата начала", "Дата окончания", "Отметка"])
    for w in works:
        ws.append([w.topic, w.position, w.start_date, w.end_date, w.completed])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="contract.xlsx"'
    wb.save(response)
    return response


@login_required
def export_scientific_excel(request):
    teacher = request.user.teacher
    works = ScientificMethodicalWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Научно-методическая"

    ws.append(["Тема", "Дата начала", "Дата окончания", "Отметка"])
    for w in works:
        ws.append([w.topic, w.start_date, w.end_date, w.completed])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="scientific.xlsx"'
    wb.save(response)
    return response


@login_required
def export_social_excel(request):
    teacher = request.user.teacher
    works = SocialEducationalWork.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Общественная работа"

    ws.append(["Наименование", "Отметка"])
    for w in works:
        ws.append([w.title, w.completed])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="social.xlsx"'
    wb.save(response)
    return response


@login_required
def export_remarks_excel(request):
    teacher = request.user.teacher
    works = TeacherRemark.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Замечания"

    ws.append(["Дата", "Содержание"])
    for w in works:
        ws.append([w.date, w.content])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="remarks.xlsx"'
    wb.save(response)
    return response


@login_required
def export_qualification_excel(request):
    teacher = request.user.teacher
    works = QualificationUpgrade.objects.filter(teacher=teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Повышение квалификации"

    ws.append(["Название", "Место", "Номер документа", "Дата", "Срок", "Объём"])
    for w in works:
        ws.append([w.title, w.location, w.document_number, w.date, w.duration, w.volume])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="qualification.xlsx"'
    wb.save(response)
    return response


from django.db.models import Model

def has_academic_year_field(model: Model):
    return 'academic_year' in [f.name for f in model._meta.fields]

def export_full_teacher_excel(request):
    teacher = request.user.teacher
    selected_year = request.GET.get("year") or request.session.get("selected_year")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(title, model_cls, columns, extract_fn):
        ws = wb.create_sheet(title=title)
        ws.append(columns)

        qs = model_cls.objects.filter(teacher=teacher)
        if selected_year and has_academic_year_field(model_cls):
            qs = qs.filter(academic_year=selected_year)

        for obj in qs:
            ws.append(extract_fn(obj))

        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    add_sheet("Уч.-методическая", EducationalMethodicalWork,
              ["Название", "Начало", "Окончание", "Отметка", "Год"],
              lambda w: [w.title, w.start_date, w.end_date, w.completed, w.academic_year])

    add_sheet("Орг.-методическая", OrganizationalMethodicalWork,
              ["Название", "Начало", "Окончание", "Отметка"],
              lambda w: [w.title, w.start_date, w.end_date, w.completed])

    add_sheet("НИР", ResearchWork,
              ["Тема", "Начало", "Окончание", "Отметка"],
              lambda w: [w.topic, w.start_date, w.end_date, w.completed])

    add_sheet("Хоздоговорная", ContractResearchWork,
              ["Тема", "Должность", "Начало", "Окончание", "Отметка"],
              lambda w: [w.topic, w.position, w.start_date, w.end_date, w.completed])

    add_sheet("Научно-методическая", ScientificMethodicalWork,
              ["Тема", "Начало", "Окончание", "Отметка"],
              lambda w: [w.topic, w.start_date, w.end_date, w.completed])

    add_sheet("Общественная работа", SocialEducationalWork,
              ["Наименование", "Отметка"],
              lambda w: [w.title, w.completed])

    add_sheet("Замечания", TeacherRemark,
              ["Дата", "Содержание"],
              lambda w: [w.date, w.content])

    add_sheet("Повышение квалификации", QualificationUpgrade,
              ["Название", "Место", "Номер документа", "Дата", "Срок", "Объём"],
              lambda w: [w.title, w.location, w.document_number, w.date, w.duration, w.volume])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'teacher_full_report_{selected_year or "all"}.xlsx'
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def teacher_profile(request):
    teacher = request.user.teacher
    return render(request, 'main/teacher_profile.html', {'teacher': teacher})

def teacher_load_view(request, teacher_id):
    teacher = Teacher.objects.get(id=teacher_id)
    disciplines = TeachingLoad.objects.filter(teacher=teacher)
    print("Преподаватель:", teacher.full_name)
    print("Найдено записей:", disciplines.count())

    return render(request, 'teacher_dashboard.html', {
        'teacher': teacher,
        'disciplines_ip': disciplines,
    })